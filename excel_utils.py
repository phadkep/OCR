import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import collections
import pandas as pd

def write_to_excel(out_array, output_excel_path):
    """
    Writes the CSV-like output to an Excel sheet.

    Args:
        out_array (numpy.ndarray): A 2D array representing the CSV-like output.
        output_excel_path (str): The path to the output Excel file.
    """
    # Load the workbook
    wb = openpyxl.load_workbook(output_excel_path)

    # Check if "Sheet2" exists, create it if it doesn't
    if 'Sheet2' not in wb.sheetnames:
        wb.create_sheet('Sheet2')

    # Select the sheet
    sheet = wb['Sheet2']

    # Convert the out_array to a DataFrame
    out_df = pd.DataFrame(out_array[1:], columns=out_array[0])

    print(out_df)

    # Write DataFrame to Excel sheet
    for r_idx, row in enumerate(dataframe_to_rows(out_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)

    # Save the workbook
    wb.save(output_excel_path)

import openpyxl

def merge_duplicate_rows(excel_file_path):
    """
    Merges duplicate rows in the Excel sheet.

    Args:
        excel_file_path (str): The path to the Excel file.
    """
    # Load the workbook
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the worksheet by name ('Sheet2' in this case)
    worksheet = workbook['Sheet2']

    # Get the first two rows
    first_row = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    second_row = list(next(worksheet.iter_rows(min_row=2, max_row=2, values_only=True)))

    # Check if the first two rows have the same text in any column
    same_text_in_any_column = any(first_row[i] == second_row[i] for i in range(len(first_row)))

    if same_text_in_any_column:
        # Merge the first two rows
        merged_row = []
        for i in range(len(first_row)):
            if first_row[i] == second_row[i]:
                merged_row.append(first_row[i])
            else:
                if first_row[i] is None or first_row[i] == '':
                    merged_row.append(second_row[i])
                elif second_row[i] is None or second_row[i] == '':
                    merged_row.append(first_row[i])
                else:
                    merged_row.append(f"{first_row[i]} {second_row[i]}")

        # Replace the first two rows with the merged row
        worksheet.delete_rows(1, 2)
        worksheet.insert_rows(1)  # Insert a new row at the beginning
        for cell, value in enumerate(merged_row, start=1):
            worksheet.cell(row=1, column=cell, value=value)
    else:
        # Keep the first two rows as they are
        pass

    # Save the modified workbook
    workbook.save(excel_file_path)

import openpyxl
from openpyxl.utils import get_column_letter

def merge_columns_with_identical_text(excel_file_path):
    """
    Merges columns with identical text in the Excel sheet.

    Args:
        excel_file_path (str): The path to the Excel file.
    """
    # Load the Excel file
    workbook = openpyxl.load_workbook(excel_file_path)

    # Select the worksheet by name ('Sheet2' in this case)
    worksheet = workbook['Sheet2']

    # Function to check if two columns have any identical text
    def have_identical_text(col1, col2):
        for cell1, cell2 in zip(col1[1:], col2[1:]):  # Start from the second row
            if cell1.value == cell2.value:
                return True
        return False

    # Function to merge columns and update headers
    def merge_columns_with_identical_text(col1, col2, col1_name, col2_name):
        merged_col_name = f"{col1_name or ''} {col2_name or ''}".strip()  # Use empty string if header is None

        for row_num in range(1, len(col1)):
            if col1[row_num].value is None:
                col1[row_num].value = col2[row_num].value

        # Delete the next column
        worksheet.delete_cols(col2[0].column)

        # Update the header
        col1[0].value = merged_col_name

    # Keep merging columns until no more duplicates are found
    while True:
        merged = False
        for col_num in range(1, worksheet.max_column):
            col1 = worksheet[get_column_letter(col_num)]
            col2 = worksheet[get_column_letter(col_num + 1)]
            col1_name = worksheet.cell(row=1, column=col_num).value
            col2_name = worksheet.cell(row=1, column=col_num + 1).value

            # Check if the current and next columns have any identical text
            if have_identical_text(col1, col2):
                merge_columns_with_identical_text(col1, col2, col1_name, col2_name)
                merged = True
                break  # Break the loop to start from the beginning after merging

        # If no more duplicates were found, exit the loop
        if not merged:
            break

    # Save the modified Excel file
    workbook.save(excel_file_path)